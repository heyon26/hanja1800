import os
import time
from openpyxl import *
from tkinter.constants import ACTIVE, COMMAND, FALSE, SOLID, TRUE
import tkinter.ttk as ttk
import tkinter.font
import tkinter as tk
from openpyxl.styles.colors import WHITE


#파이썬 실행 경로설정
#os.chdir('C:\Users\admin\Downloads\hanja1800-main')

#openpyxl 파일/시트 불러오기
load_wb = load_workbook("1800ff.xlsx", data_only=True)
load_ws = load_wb['1800']

#tkinter 기본설정. 제목. 창크기설정
root = tk.Tk()
root.title("한자")
root.geometry("400x200+0+0")
root.configure(bg='white')
root.attributes('-topmost', 'true')

#폰트 설정
font1=tkinter.font.Font(family="바탕", size=20)
font=tkinter.font.Font(family="바탕", size=60)
font2=tkinter.font.Font(family="바탕", size=28)
font3=tkinter.font.Font(family="바탕", size=20)


#콤보박스: 스트링바=텍스트를 바꿔야할 때 사용/ 기본설정/선택값설정/콤보박스이름설정/초기값설정/위치설정
str=tk.StringVar()
combobox = ttk.Combobox(height=0, textvariable=str)
combobox['values']=('1-200', '201-400', '401-600', '601-800', '801-1000', '1001-1200', '1201-1400', '1401-1600', '1601-1800')
combobox.set("정렬")
combobox.current=(0)
combobox.grid(row=0,column=2)

strs=tk.StringVar()
comboboxs = ttk.Combobox(height=0, textvariable=strs, width=5)
comboboxs['values']=('0-1','0-2','0-3','1-1','1-2','1-3','2-1','2-2','2-3')
comboboxs.set("초설정")
comboboxs.current=(0)
comboboxs.grid(row=0,column=0)

#숫자출력라벨: 스트링바/초기값/기본값(위치, textvariable=변경되는텍스트값=변수, 폰트)/위치설정/위치설정2(넓이값, y좌표값)
label1 = tk.StringVar()
label1.set("0")
reallabel1 = tk.Label(anchor='s', textvariable=label1, font=font1, bg='white')
reallabel1.grid(row=1, column=0)
reallabel1.place(width=80, rely=0.45)

label = tk.StringVar()
label.set("漢")                                                   #배경선 굵기/  배경색    /선 옵션
reallabel = tk.Label(anchor='center', textvariable=label, font=font, bg='white')
reallabel.grid(row=1, column=1)
reallabel.place(width=75, rely=0.2, relx=0.20)

label3 = tk.StringVar()
label3.set("뜻")
reallabel3 = tk.Label(anchor='w', textvariable=label3, font=font3, bg='white')
reallabel3.grid(row=1, column=2)
reallabel3.place(width=180, rely=0.45, relx=0.45)

label4 = tk.StringVar()
label4.set("부수")
reallabel4 = tk.Label(anchor='w', textvariable=label4, font=font3, bg='white')
reallabel4.grid(row=2, column=1)
reallabel4.place(width=180, rely=0.8, relx=0.25)

label5 = tk.StringVar()
label5.set("뜻")
reallabel5 = tk.Label(anchor='w', textvariable=label5, font=font3, bg='white')
reallabel5.grid(row=2, column=2)
reallabel5.place(width=180, rely=0.8, relx=0.45)


#중지를 위한 변수
k=1
#콤보박스 값 받아온 변수
cbv=str.get()
cbv2=strs.get()
#def bttn에 들어가는 range용 변수 /d =역순용
b=int(0)
c=int(0)
d=int(1)
#초설정용
bs1=int(1)
bs2=int(2)
#유사for문용
bttni=int(0)
#combobox 값을 받아서 split한뒤 int로 바꿈
def number():
    global cbv,b,c,d,k,bttni #global k 문장은 함수 안에서 함수 밖의 k 변수를 직접 사용
    a= str.get().split("-")
    b= int(a[0])
    c= int(a[1])
    if b>c:
        d=-1
        bttni=b+2
    else:
        d=1
        bttni=b
    k= 1
    print(b,type(b),c,type(c), d)
    cl.set(str.get())
    root.update()

def setTime():
    global bs1,bs2,cbv2
    a= strs.get().split("-")
    bs1= int(a[0])
    bs2= int(a[1])

#b = 앞숫자 c=뒤 d= 앞뒤
# 실행용
def bttn():
    global b,c,d,bttni
    if k!=1:
        print("stop")
    else:
        if d==1:
            if bttni == c+1:
                bttni = b+1
                root.after(0,bttnF)
            else:
                bttni += 1
                root.after(0,bttnF)
        elif d==-1:
            if bttni == c+1:
                bttni = b+1
                root.after(0,bttnF)
            else:
                print(bttni,c)
                bttni -= 1
                root.after(0,bttnF)



def bttnF():
    global bttni
    label1.set(bttni-1) #번호
    label.set(load_ws.cell(row=bttni, column=2).value) #한자
    label4.set(load_ws.cell(row=bttni, column=4).value)
    label5.set(" ")
    label3.set(" ")
    root.update()
    root.after(bs1*1000, bttnS) 

def bttnS():
    label3.set(load_ws.cell(row=bttni, column=3).value) #뜻음
    label5.set(load_ws.cell(row=bttni, column=5).value)
    root.update() #update 해야 반영됨
    root.after(bs2*1000, bttn)

# 중지용
def bttn2():
    global k,bttni  #global k 문장은 함수 안에서 함수 밖의 k 변수를 직접 사용
    k=0
    bttni=b
    label1.set("0")
    label.set("漢")
    label3.set("stop")
    label4.set(" ")
    label5.set(" ")
    root.update()


btns=tk.Button(text="sec", command=setTime)
btns.grid(row=0, column=1)   
btn0 =tk.Button(text="set", command=number)#command=number()을 쓸 경우 안눌러도 호출해버림
btn0.grid(row=0, column=3)
btn =tk.Button(text="start", command=bttn)
btn.grid(row=0, column=4)
btn2 =tk.Button(text="stop", command=bttn2)
btn2.grid(row=0, column=5)
cl = tk.StringVar()
cl.set(str.get())
checklabel=tk.Label(textvariable=cl)
checklabel.grid(row=0, column=6)

root.mainloop()

#for문 속에서는 after(sec) 때문에 창 이동, 다른 def 실행할 시 딜레이 발생/after(sec,func)쓸 수도 없음
#if문으로 for문 구현해서 순차실행.